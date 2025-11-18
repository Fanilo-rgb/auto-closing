import openpyxl as xl
from io import BytesIO

def create_closing_excel(data):
    wb = xl.load_workbook('for use.xlsx')
    order_summary = wb['ORDER SUMMARY']
    application_summary = wb['APPLICATON SUMAMRY']
    order_summary_starting_row = 5
    application_summary_starting_row = 4

    for person in data["bvs"]:
        card_cell = order_summary.cell(order_summary_starting_row, 2)
        name_cell = order_summary.cell(order_summary_starting_row, 3)
        card_cell.value = person['numberCard']
        name_cell.value = person['name']

        products = person['products']

        for product in products:
            cell = order_summary.cell(order_summary_starting_row, product['order'] + 8)
            cell.value = product['quantity']

        order_summary_starting_row += 1

    for person in data["applications"]:
        new_person_card_cell = application_summary.cell(application_summary_starting_row, 2)
        new_person_name_cell = application_summary.cell(application_summary_starting_row, 3)
        new_person_phone_cell = application_summary.cell(application_summary_starting_row, 10)
        new_person_cin_cell = application_summary.cell(application_summary_starting_row, 11)

        sponsor_card_cell = application_summary.cell(application_summary_starting_row, 4)
        sponsor_name_cell = application_summary.cell(application_summary_starting_row, 5)

        up_line_card_cell = application_summary.cell(application_summary_starting_row, 6)
        up_line_name_cell = application_summary.cell(application_summary_starting_row, 7)


        new_person_card_cell.value = person["consultant"]["numberCard"]
        new_person_name_cell.value = person["consultant"]["name"]
        new_person_phone_cell.value = person["consultant"]["phone"]
        new_person_cin_cell.value = person["consultant"]["cin"]

        sponsor_card_cell.value = person["sponsor"]["numberCard"]
        sponsor_name_cell.value = person["sponsor"]["name"]

        up_line_card_cell.value = person["upLine"]["numberCard"]
        up_line_name_cell.value = person["upLine"]["name"]

        application_summary_starting_row += 1


    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output
